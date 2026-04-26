import time
from openai import OpenAI
from openpyxl import load_workbook

GROQ_API_KEY = "GROQ_API_KEY"
OPENROUTER_API_KEY = "OPENROUTER_API_KEY"

groq_client = OpenAI(api_key=GROQ_API_KEY, base_url="https://api.groq.com/openai/v1")
openrouter_client = OpenAI(api_key=OPENROUTER_API_KEY, base_url="https://openrouter.ai/api/v1")

MODELS = [
    ("DeepSeek-R1 Prediction", openrouter_client, "deepseek/deepseek-r1"),
    ("Llama 4 Prediction", groq_client, "meta-llama/llama-4-scout-17b-16e-instruct"),
    ("Qwen 2.5 Prediction", openrouter_client, "qwen/qwen-2.5-72b-instruct"),
]

def classify_email(client, model, subject, body):
    prompt = (
        "You are a cybersecurity email analyst. "
        "Analyze the following email and determine whether it is a phishing attempt or legitimate.\n\n"
        "Subject: " + str(subject) + "\n"
        "Email Body:\n" + str(body) + "\n\n"
        "Respond in this exact format:\n"
        "Classification: [PHISHING or LEGITIMATE]\n"
        "Reasoning: [Brief explanation]"
    )

    for attempt in range(3):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=300,
                temperature=0
            )

            result = response.choices[0].message.content.strip()

            if "PHISHING" in result.upper():
                return "Phishing", result
            elif "LEGITIMATE" in result.upper():
                return "Legitimate", result
            else:
                return "Unclear", result

        except Exception as e:
            print("Attempt " + str(attempt + 1) + " failed: " + str(e))
            time.sleep(5)

    return "Error", "Failed after 3 attempts"


def process_file(filepath, output_path):
    print("\nProcessing: " + filepath)

    wb = load_workbook(filepath)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers)}

    subject_col = col.get("Subject")
    body_col = col.get("Email Body")
    notes_col = col.get("Notes")

    total = ws.max_row - 1

    for row in range(2, ws.max_row + 1):
        subject = ws.cell(row=row, column=subject_col).value or ""
        body = ws.cell(row=row, column=body_col).value or ""
        email_id = ws.cell(row=row, column=1).value

        print("[" + str(row - 1) + "/" + str(total) + "] Email ID " + str(email_id))

        notes_parts = []

        for col_name, client, model in MODELS:
            print(" -> " + col_name + "...")

            classification, full_response = classify_email(client, model, subject, body)

            ws.cell(row=row, column=col.get(col_name), value=classification)
            notes_parts.append(col_name + ": " + full_response)

            time.sleep(2)

        ws.cell(row=row, column=notes_col, value="\n\n".join(notes_parts))

        wb.save(output_path)

    print("Saved: " + output_path)


base = "data/"

files = [
    (base + "\\1 Legitimate Emails.xlsx", base + "\\1 Legitimate Emails_RESULTS.xlsx"),
    (base + "\\2 AI Generated Phishing Emails.xlsx", base + "\\2 AI Generated Phishing Emails_RESULTS.xlsx"),
    (base + "\\3 Human Crafted Phishing Emails.xlsx", base + "\\3 Human Crafted Phishing Emails_RESULTS.xlsx"),
]

for input_path, output_path in files:
    process_file(input_path, output_path)

print("\nALL DONE! Check your RESULTS files in the CAPSTONE folder.")
