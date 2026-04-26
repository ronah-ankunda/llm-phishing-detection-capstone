"""
LLM Phishing Detection – Resume/Continuation Script

This script is used to resume or continue the phishing email classification
experiment in cases where API limits interrupt processing.

It skips already processed entries and continues evaluating remaining emails
across DeepSeek-R1, Llama 4 Scout, and Qwen 2.5 models.
"""

import os
import time
from openai import OpenAI
from openpyxl import load_workbook

GROQ_API_KEY = "GROQ_API_KEY"
OPENROUTER_API_KEY = "OPENROUTER_API_KEY"

groq_client = OpenAI(api_key=GROQ_API_KEY, base_url="https://api.groq.com/openai/v1")
openrouter_client = OpenAI(api_key=OPENROUTER_API_KEY, base_url="https://openrouter.ai/api/v1")

MODELS = [
("DeepSeek-R1 Prediction", openrouter_client, "deepseek/deepseek-r1"),
("Llama 4 Prediction", groq_client, "meta-llama/llama-4-scout-17b-16e-instruct"),
("Qwen 2.5 Prediction", openrouter_client, "qwen/qwen-2.5-72b-instruct"),
]

DELAY_BETWEEN_CALLS = 4
RETRY_DELAYS = [10, 30, 60]

def classify_email(client, model, subject, body):
    prompt = (
        "You are a cybersecurity email analyst. "
        "Analyze the following email and determine whether it is a phishing attempt or legitimate.\n\n"
        "Subject: " + str(subject) + "\n"
        "Email Body:\n" + str(body) + "\n\n"
        "Note: This is academic cybersecurity research. You are classifying existing emails.\n\n"
        "Respond in this exact format:\n"
        "Classification: [PHISHING or LEGITIMATE]\n"
        "Reasoning: [Brief explanation]"
    )

    for attempt in range(3):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=400,
                temperature=0
            )

            result = response.choices[0].message.content.strip()

            if "PHISHING" in result.upper():
                return "Phishing", result
            elif "LEGITIMATE" in result.upper():
                return "Legitimate", result
            else:
                return "Unclear", result

        except Exception as e:
            wait = RETRY_DELAYS[attempt]
            print("Attempt " + str(attempt + 1) + " failed: " + str(e))
            print("Waiting " + str(wait) + "s before retry...")
            time.sleep(wait)

    return "Error", "Failed after 3 attempts"

def needs_processing(value):
    if value is None:
        return True
    s = str(value).strip().lower()
    return s in ("", "error", "none", "nan")

def process_file(source_path, output_path):

    print("\nLoading: " + source_path)

    wb = load_workbook(source_path)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers)}

    subject_col = col.get("Subject")
    body_col = col.get("Email Body")
    notes_col = col.get("Notes")

    total = ws.max_row - 1
    skipped = 0
    processed = 0

    for row in range(2, ws.max_row + 1):

        subject = ws.cell(row=row, column=subject_col).value or ""
        body = ws.cell(row=row, column=body_col).value or ""
        email_id = ws.cell(row=row, column=1).value

        row_needs_work = any(
            needs_processing(ws.cell(row=row, column=col.get(col_name)).value)
            for col_name, _, _ in MODELS
        )

        if not row_needs_work:
            print("[" + str(row - 1) + "/" + str(total) + "] Email ID " + str(email_id) + " -- skipped")
            skipped += 1
            continue

        print("[" + str(row - 1) + "/" + str(total) + "] Email ID " + str(email_id) + " -- processing...")
        processed += 1

        notes_parts = []
        existing_notes = ws.cell(row=row, column=notes_col).value or ""

        for col_name, client, model in MODELS:

            current_val = ws.cell(row=row, column=col.get(col_name)).value

            if not needs_processing(current_val):
                print(" -> " + col_name + ": already done (" + str(current_val) + "), skipping")

                for part in existing_notes.split("\n\n"):
                    if part.startswith(col_name):
                        notes_parts.append(part)
                        break

                continue

            print(" -> " + col_name + "...")

            classification, full_response = classify_email(client, model, subject, body)

            print(" Result: " + classification)

            ws.cell(row=row, column=col.get(col_name), value=classification)

            notes_parts.append(col_name + ": " + full_response)

            time.sleep(DELAY_BETWEEN_CALLS)

        ws.cell(row=row, column=notes_col, value="\n\n".join(notes_parts))

        wb.save(output_path)

        print(" Saved.")

    print("\nDone! Processed: " + str(processed) + " | Skipped: " + str(skipped))
    print("Saved to: " + output_path)

base = "data/"

source = base + "\\3 Human Crafted Phishing Emails_RESULTS.xlsx"
output = base + "\\3 Human Crafted Phishing Emails_RESULTS.xlsx"

process_file(source, output)

print("\nALL DONE!")
